#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time : 2021-04-01 11:19
# @Author : Pan
# @Version：V 0.1
# @File : main.py
# @desc : 解析电子发票信息，将内容写入excel台账

import base64
import json
import os
import shutil
import time
import urllib
import urllib.parse
import urllib.request

import fitz
import openpyxl

# 待处理发票所在目录，可根据实际情况自行修改
scr_path = r"E:\桌面\新建文件夹"
dst_path = os.path.join(scr_path, '解析成功文件')
info_list_cn = ['序号', '发票种类', '发票名称', '发票代码', '发票号码', '校验码', '开票日期', '购方名称', '购方纳税人识别号',
                '购方地址及电话', '购方开户行及账号', '货物名称', '规格型号', '单位', '数量', '单价', '金额', '税率', '税额',
                '合计金额', '合计税额', '价税合计(大写)', '价税合计(小写)', '销售方名称', '销售方纳税人识别号', '销售方地址及电话',
                '销售方开户行及账号', '备注']
info_list_en = ['ID', 'InvoiceType', 'InvoiceTypeOrg', 'InvoiceCode', 'InvoiceNum', 'CheckCode', 'InvoiceDate',
                'PurchaserName', 'PurchaserRegisterNum', 'PurchaserAddress', 'PurchaserBank', 'CommodityName',
                'CommodityType', 'CommodityUnit', 'CommodityNum', 'CommodityPrice', 'CommodityAmount',
                'CommodityTaxRate',
                'CommodityTax', 'TotalAmount', 'TotalTax', 'AmountInWords', 'AmountInFiguers', 'SellerName',
                'SellerRegisterNum', 'SellerAddress', 'SellerBank', 'Remarks']


# 检查文件夹内的文件
def check_data():
    """
    @return check_files:返回符合条件的pdf或ofd电子发票文件
    """
    if not os.path.exists(scr_path):
        os.makedirs(scr_path)  # 创建目录或文件夹
        print(f"您还没有目录{scr_path},刚才已自动为您创建完毕,\n请将待处理电子发票文件或文件夹存入该目录下并重新运行程序。")
        exit()
    if not os.path.exists(dst_path):
        os.makedirs(dst_path)
    check_files = []
    for file in os.walk(scr_path):
        for f in file[2]:
            path = os.path.join(file[0], f)
            if (os.path.isfile(path) and os.path.splitext(path)[1] == '.pdf' or os.path.isfile(path) and
                    os.path.splitext(path)[1] == '.ofd'):
                check_files.append(path)
    print("总计有" + str(len(check_files)) + '个pdf和ofd文件')
    if len(check_files) < 1:
        print('文件夹中没有可识别的pdf或ofd文件,请检查后重新执行程序')
        exit()
    return check_files


# 获取token
# noinspection PyUnboundLocalVariable
def get_token():
    # 百度AI账号获取地址 https://ai.baidu.com/tech/ocr_receipts/vat_invoice
    # client_id 为官网获取的AK， client_secret 为官网获取的SK
    client_id = 'WhNRQgE0HPzFYiBuvsKUrYQH'
    client_secret = 'fqQLUxvkdey1HX0H7BMIPMaK5ecz3yOw'

    host = 'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=' + client_id + '&client_secret=' + client_secret
    request = urllib.request.Request(host)
    request.add_header('Content-Type', 'application/json; charset=UTF-8')
    response = urllib.request.urlopen(request)
    token_content = response.read()
    if token_content:
        token_info = json.loads(token_content)
        token_key = token_info['access_token']
    return token_key


# 调用百度接口识别电子发票内容
def vat_invoice(filename, a):
    """
    @param filename:要解析的文件名称
    @return invoice_list:解析出的电子发票信息
    """
    request_url = "https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice"
    fl = open(filename, 'rb')
    img = base64.b64encode(fl.read())
    params = dict()
    params['image'] = img
    params['show'] = 'true'
    params = urllib.parse.urlencode(params).encode("utf-8")
    access_token = get_token()
    request_url = request_url + "?access_token=" + access_token
    request = urllib.request.Request(url=request_url, data=params)
    request.add_header('Content-Type', 'application/x-www-form-urlencoded')
    response = urllib.request.urlopen(request)
    content = response.read()
    invoice_list = []
    if content:
        content = content.decode('utf-8')
        data = json.loads(content)
        words_result = data['words_result']
        text = words_result
        for k in range(len(text['CommodityName'])):
            list1 = []
            for i in info_list_en:
                try:
                    if k == 0:
                        if i == 'ID':
                            list1.append(a)
                        elif i.startswith('Comm'):
                            list1.append(text[i][k]['word'])
                        else:
                            list1.append(text[i])
                    else:
                        if i == 'ID':
                            list1.append('')
                        elif i.startswith('Comm'):
                            list1.append(text[i][k]['word'])
                        else:
                            list1.append('')
                except:
                    list1.append('')
            invoice_list.append(list1)
    time.sleep(0.5)
    fl.close()
    print('解析完毕: ' + filename)
    return invoice_list


# 使用openpyxl写入excel
def write_to_excel(path: str, sheetStr, info, data):
    """
    @param path: excel保存路径
    @param sheetStr: 工作表名称
    @param info: 表头信息list
    @param data: 写入数据，二维列表
    @return: 无
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # 为sheet设置一个title
    ws.title = sheetStr
    # 添加表头（不需要表头可以不用加）
    data.insert(0, list(info))
    # 开始遍历数组
    for row_index, row_item in enumerate(data):
        for col_index, col_item in enumerate(row_item):
            ws.cell(row=row_index + 1, column=col_index + 1, value=col_item)

    # 设置表格列格式
    columns1 = ['B', 'C', 'E']
    columns2 = ['T', 'U', 'V', 'H', 'I', 'L']
    ws.column_dimensions['A'].width = 4.5
    for column1 in columns1:
        ws.column_dimensions[column1].width = 14
    for column2 in columns2:
        ws.column_dimensions[column2].width = 20

    # 写入excel文件 如果path路径的文件不存在那么就会自动创建
    wb.save(path)
    print('写入成功')


# 判断重复票
def identify_same_row(file, ws):
    """
    @param file: 传入一个excel文件的绝对路径字符串
    @param ws: 工作表名称
    @return: 无
    """
    mylist = []
    irow = ws.max_row
    for row in range(2, irow + 1):
        try:
            code1 = ws.cell(row, 3).value + ws.cell(row, 4).value
        except:
            code1 = ''
        mylist.append(code1)
    dub = {}
    for code in mylist:
        if mylist.count(code) > 1:
            dub[code] = mylist.count(code)
    for r in range(2, irow + 1):
        try:
            code2 = ws.cell(r, 3).value + ws.cell(r, 4).value
        except:
            code2 = ''
        ws.cell(r, 15).value = dub.get(code2)
        if ws.cell(r, 15).value is not None:
            ws.cell(r, 15).fill = fill_du  # 设置重复内容浅黄色标注
            ws.cell(r, 14).fill = fill_du  # 设置重复内容浅黄色标注


def main(files_list):
    """
    @param files_list:传入文件列表
    """
    a = 0
    results_list = []
    for file in files_list:
        fname = files_list[a]
        a = a + 1
        if fname[-4:] == '.pdf':
            with fitz.open(file) as pdf:
                for pg in range(pdf.pageCount):
                    page = pdf[pg]
                    rotate = int(0)
                    zoom_x = 2.0
                    zoom_y = 2.0
                    trans = fitz.Matrix(zoom_x, zoom_y).preRotate(rotate)
                    pm = page.getPixmap(matrix=trans, alpha=False)
                    image = fname[0:-4] + '.png'
                    pm.writePNG(image)
                    pg_text = page.seachFor('销售货物或者提供应税劳务、服务清单')
                    try:
                        list11 = vat_invoice(image, a)
                        for ml in list11:
                            results_list.append(ml)
                    except Exception as e:
                        print('执行错误：', e, file)  # 删除作业过程中产生的png文件
                        continue
                    os.remove(image)  # 删除作业过程中产生的png文件

            new_name = f'{list11[0][7]}-{list11[0][4]}-{list11[0][6]}-电子发票.pdf'
            new_path = os.path.join(dst_path, new_name)
            shutil.move(file, new_path)
    print('共计发现：' + str(len(files_list)) + '个文件，处理了' + str(a) + '个')
    return results_list


if __name__ == '__main__':
    files = check_data()
    data_list = main(files)
    t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    save_path = os.path.join(dst_path, f'{t}-电子发票管理台账.xlsx')
    write_to_excel(path=save_path, sheetStr='1', info=info_list_cn, data=data_list)
    print('处理结束！')
